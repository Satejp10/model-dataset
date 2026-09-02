#!/usr/bin/env python3
"""
diff_dataset.py — diff a fresh Models Table export against the last snapshot,
then log the delta.

The Models Table exports (CSV / JSON / XLSX) are subscriber-gated, so this
script does NOT fetch anything over the network. You supply a fresh export; it
does the diff + logging.

Typical use
-----------
1. Download a fresh export from the Models Table and save it over
   "Pruned AI Models_Table.xlsx" in the repo root (or pass --new PATH to point
   at a file somewhere else).
2. Preview the delta (writes nothing):
       python logs/diff_dataset.py
3. Record it:
       python logs/diff_dataset.py --write
   which will
     - prepend a dated entry to logs/CHANGELOG.md,
     - refresh the "Current dataset specs" block in logs/CHANGELOG.md,
     - save the new baseline to logs/snapshots/latest.json.

On the very first run there is no baseline yet, so --write just seeds
logs/snapshots/latest.json from the current file and records no diff.

Only dependency: openpyxl  (pip install openpyxl)
"""

from __future__ import annotations

import argparse
import datetime as _dt
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ModuleNotFoundError:
    sys.exit("openpyxl is required: pip install openpyxl")

LOGS_DIR = Path(__file__).resolve().parent
REPO_ROOT = LOGS_DIR.parent
DEFAULT_XLSX = REPO_ROOT / "Pruned AI Models_Table.xlsx"
SNAPSHOT = LOGS_DIR / "snapshots" / "latest.json"
CHANGELOG = LOGS_DIR / "CHANGELOG.md"

SHEET_NAME = "Models"
HEADER_ROW = 2  # the real header row; row 1 holds permalinks/metadata

SPECS_START = "<!-- SPECS:START -->"
SPECS_END = "<!-- SPECS:END -->"
CHANGES_START = "<!-- CHANGES:START -->"

# Labels that refer to the same organisation, shown as a note in the specs block.
LAB_MERGE_NOTE = (
    "> Note: `Google DeepMind`, `Google`, and `DeepMind` are recorded as separate\n"
    "> labels. If treated as one organisation, subtract 2 from the distinct-lab count."
)


# --------------------------------------------------------------------------- #
# Parsing
# --------------------------------------------------------------------------- #
def _clean_header(name) -> str:
    """Tidy a header cell: drop sort glyphs, collapse embedded newlines/spaces."""
    s = str(name).replace("​", "")
    s = re.sub(r"[▼▲△▽]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return re.sub(r"(\w) -(\w)", r"\1-\2", s)  # rejoin a hyphenated word split across lines


def _norm(value) -> str:
    """Stable string form of a cell value so equal cells compare equal."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.date().isoformat() if isinstance(value, _dt.datetime) else value.isoformat()
    return str(value).strip()


def load_models(path: Path, sheet_name: str = SHEET_NAME, header_row: int = HEADER_ROW) -> dict:
    """Return {columns, models, order} parsed from an export file."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < header_row:
        sys.exit(f"{path.name}: sheet '{sheet_name}' has no header at row {header_row}")

    header = rows[header_row - 1]
    # keep only named columns, remembering their positional index (skips the blank spacer)
    cols = [(i, _clean_header(h)) for i, h in enumerate(header) if h is not None and _clean_header(h)]
    col_names = [name for _, name in cols]
    if "Model" not in col_names:
        sys.exit(f"{path.name}: could not find a 'Model' column in header row {header_row}")

    model_idx = next(i for i, name in cols if name == "Model")
    lab_idx = next((i for i, name in cols if name == "Lab"), None)

    # first pass: which model names collide (need lab-qualified keys)
    names: dict[str, int] = {}
    for r in rows[header_row:]:
        m = _norm(r[model_idx]) if model_idx < len(r) else ""
        if m:
            names[m] = names.get(m, 0) + 1

    models: dict[str, dict] = {}
    order: list[str] = []
    for r in rows[header_row:]:
        model = _norm(r[model_idx]) if model_idx < len(r) else ""
        if not model:
            continue  # trailing/blank row
        lab = _norm(r[lab_idx]) if lab_idx is not None and lab_idx < len(r) else ""
        key = f"{model} ‹{lab}›" if names[model] > 1 else model
        record = {name: (_norm(r[i]) if i < len(r) else "") for i, name in cols}
        models[key] = record
        order.append(key)

    return {"columns": col_names, "models": models, "order": order}


def compute_specs(parsed: dict) -> dict:
    labs: dict[str, int] = {}
    for rec in parsed["models"].values():
        lab = rec.get("Lab", "") or "(blank)"
        labs[lab] = labs.get(lab, 0) + 1
    per_lab = dict(sorted(labs.items(), key=lambda kv: (-kv[1], kv[0])))
    return {
        "num_models": len(parsed["models"]),
        "num_labs": len(per_lab),
        "num_columns": len(parsed["columns"]),
        "columns": parsed["columns"],
        "per_lab": per_lab,
    }


# --------------------------------------------------------------------------- #
# Diff
# --------------------------------------------------------------------------- #
def diff_models(old: dict, new: dict, new_order: list[str]):
    old_keys, new_keys = set(old), set(new)
    added = [k for k in new_order if k in new_keys - old_keys]
    removed = sorted(old_keys - new_keys)
    changed: dict[str, dict] = {}
    for k in new_order:
        if k not in old_keys:
            continue
        o, n = old[k], new[k]
        fields = {}
        for col, nv in n.items():
            ov = o.get(col, "")
            if ov != nv:
                fields[col] = (ov, nv)
        for col, ov in o.items():
            if col not in n and ov != "":
                fields[col] = (ov, "")
        if fields:
            changed[k] = fields
    return added, removed, changed


def _label(key: str, models: dict) -> str:
    rec = models.get(key, {})
    model = rec.get("Model", key)
    lab = rec.get("Lab", "")
    return f"{model} — {lab}" if lab else str(model)


# --------------------------------------------------------------------------- #
# Rendering
# --------------------------------------------------------------------------- #
def render_console(added, removed, changed, old_specs, new_specs, new_models, old_models) -> str:
    out = []
    out.append(f"models  {old_specs['num_models']:>4} -> {new_specs['num_models']:<4}"
               f"   labs {old_specs['num_labs']:>3} -> {new_specs['num_labs']:<3}"
               f"   columns {old_specs['num_columns']:>3} -> {new_specs['num_columns']:<3}")
    out.append(f"added {len(added)}   removed {len(removed)}   changed {len(changed)}")
    out.append("")
    if added:
        out.append("ADDED")
        out += [f"  + {_label(k, new_models)}" for k in added]
    if removed:
        out.append("REMOVED")
        out += [f"  - {_label(k, old_models)}" for k in removed]
    if changed:
        out.append("CHANGED")
        for k in changed:
            out.append(f"  ~ {_label(k, new_models)}")
            for col, (ov, nv) in changed[k].items():
                out.append(f"      {col}: {ov!r} -> {nv!r}")
    if not (added or removed or changed):
        out.append("No changes.")
    return "\n".join(out)


def render_changelog_entry(date, source_name, added, removed, changed,
                           old_specs, new_specs, new_models, old_models) -> str:
    lines = [f"### {date} — Update (+{len(added)} / -{len(removed)} / ~{len(changed)})", ""]
    lines.append(
        f"Source export: `{source_name}` · "
        f"models {old_specs['num_models']} → {new_specs['num_models']} · "
        f"labs {old_specs['num_labs']} → {new_specs['num_labs']} · "
        f"columns {old_specs['num_columns']} → {new_specs['num_columns']}"
    )
    lines.append("")

    lines.append(f"**Models added ({len(added)}):**")
    lines += [f"- {_label(k, new_models)}" for k in added] or ["- none"]
    lines.append("")

    lines.append(f"**Models removed ({len(removed)}):**")
    lines += [f"- {_label(k, old_models)}" for k in removed] or ["- none"]
    lines.append("")

    lines.append(f"**Models changed ({len(changed)}):**")
    if changed:
        for k in changed:
            lines.append(f"- {_label(k, new_models)}")
            for col, (ov, nv) in changed[k].items():
                lines.append(f"  - {col}: `{ov or '∅'}` → `{nv or '∅'}`")
    else:
        lines.append("- none")
    return "\n".join(lines).rstrip() + "\n"


def render_specs_block(date, source_name, specs) -> str:
    lines = [f"_Snapshot as of {date}_", ""]
    lines += [
        "| Property | Value |",
        "|---|---|",
        f"| File | `{source_name}` |",
        f"| Sheet | `{SHEET_NAME}` |",
        f"| Header row | Row {HEADER_ROW} |",
        f"| Number of models | **{specs['num_models']}** |",
        f"| Number of distinct labs (as-written) | **{specs['num_labs']}** |",
        f"| Columns (named) | **{specs['num_columns']}** |",
        "",
        "### Columns",
        "",
        "| # | Column |",
        "|---|---|",
    ]
    lines += [f"| {i} | {c} |" for i, c in enumerate(specs["columns"], 1)]
    lines += ["", "### Models per lab", "", "| Lab | Models |", "|---|---:|"]
    lines += [f"| {lab} | {n} |" for lab, n in specs["per_lab"].items()]
    lines += ["", LAB_MERGE_NOTE]
    return "\n".join(lines)


# --------------------------------------------------------------------------- #
# File updates
# --------------------------------------------------------------------------- #
def _replace_between(text, start, end, block) -> str:
    if start not in text or end not in text:
        sys.exit(f"CHANGELOG.md is missing the {start} / {end} markers.")
    s = text.index(start) + len(start)
    e = text.index(end)
    return text[:s] + "\n" + block + "\n" + text[e:]


def _prepend_after(text, marker, block) -> str:
    if marker not in text:
        sys.exit(f"CHANGELOG.md is missing the {marker} marker.")
    i = text.index(marker) + len(marker)
    return text[:i] + "\n\n" + block.rstrip() + "\n" + text[i:]


def save_snapshot(parsed, specs, date, source_name):
    SNAPSHOT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "captured": date,
        "source_file": source_name,
        "specs": specs,
        "columns": parsed["columns"],
        "order": parsed["order"],
        "models": parsed["models"],
    }
    SNAPSHOT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--new", type=Path, default=DEFAULT_XLSX,
                    help=f"fresh export to diff (default: {DEFAULT_XLSX.name})")
    ap.add_argument("--write", action="store_true",
                    help="apply the delta to CHANGELOG.md and save the new baseline snapshot")
    ap.add_argument("--date", default=_dt.date.today().isoformat(),
                    help="date stamp for the entry/snapshot (default: today)")
    args = ap.parse_args()

    if not args.new.exists():
        sys.exit(f"Export not found: {args.new}")

    parsed = load_models(args.new)
    specs = compute_specs(parsed)
    source_name = args.new.name

    # -------- first run: seed the baseline -------- #
    if not SNAPSHOT.exists():
        print(f"No baseline snapshot at {SNAPSHOT.relative_to(REPO_ROOT)}.")
        print(f"Parsed {specs['num_models']} models, {specs['num_labs']} labs, "
              f"{specs['num_columns']} columns from {source_name}.")
        if not args.write:
            print("\nRun again with --write to seed the baseline.")
            return
        save_snapshot(parsed, specs, args.date, source_name)
        text = CHANGELOG.read_text(encoding="utf-8")
        text = _replace_between(text, SPECS_START, SPECS_END,
                                render_specs_block(args.date, source_name, specs))
        CHANGELOG.write_text(text, encoding="utf-8")
        print(f"Seeded baseline snapshot and refreshed the specs block ({args.date}).")
        return

    # -------- normal run: diff against baseline -------- #
    old = json.loads(SNAPSHOT.read_text(encoding="utf-8"))
    old_models = old.get("models", {})
    old_specs = old.get("specs") or compute_specs({"columns": old.get("columns", []), "models": old_models})

    added, removed, changed = diff_models(old_models, parsed["models"], parsed["order"])
    print(render_console(added, removed, changed, old_specs, specs, parsed["models"], old_models))

    if not (added or removed or changed):
        print("\nNothing to log.")
        return

    if not args.write:
        print("\nPreview only. Re-run with --write to record this in CHANGELOG.md.")
        return

    entry = render_changelog_entry(args.date, source_name, added, removed, changed,
                                   old_specs, specs, parsed["models"], old_models)
    text = CHANGELOG.read_text(encoding="utf-8")
    text = _replace_between(text, SPECS_START, SPECS_END,
                            render_specs_block(args.date, source_name, specs))
    text = _prepend_after(text, CHANGES_START, entry)
    CHANGELOG.write_text(text, encoding="utf-8")
    save_snapshot(parsed, specs, args.date, source_name)
    print(f"\nLogged to {CHANGELOG.relative_to(REPO_ROOT)} and updated the baseline snapshot.")


if __name__ == "__main__":
    main()
